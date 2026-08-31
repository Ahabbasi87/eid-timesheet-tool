python
"""
Populates supplier timesheet templates from matched employees.

Loads each template with formatting intact, writes into a COPY (never
the original), and only ever writes into the configured data
columns/rows - it never touches the template's existing styling,
merged cells, or formulas elsewhere on the sheet.

Only EID No, Name, and Designation are written - all other columns
(time in/out, break, total hours, etc.) are left exactly as the
template already has them.
"""
import re
import shutil
from pathlib import Path
from typing import Dict, List

import openpyxl

from app.core.config import TIMESHEET_COLUMNS, TIMESHEET_FIRST_DATA_ROW
from app.models.schemas import LogEntry, LogSeverity, MatchResult, MatchStatus


def normalize_supplier(name: str) -> str:
    """
    Case/whitespace-insensitive key for matching a master-data supplier
    name against a template's H5 supplier name. Real-world files are
    rarely byte-identical (extra spaces, different case, punctuation),
    so matching on the raw string causes a template to silently get
    zero employees written into it.
    """
    if not name:
        return ""
    return re.sub(r"\s+", " ", name).strip().casefold()


def read_supplier_name(filepath: str) -> str:
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active
    from app.core.config import SUPPLIER_NAME_CELL
    value = ws[SUPPLIER_NAME_CELL].value
    wb.close()
    return str(value).strip() if value else ""


def populate_supplier_timesheet(
    template_path: str,
    output_path: str,
    matched_results: List[MatchResult],
) -> List[LogEntry]:
    """
    Writes matched employees' EID No, Name, and Designation into a copy
    of the template. Every other cell in the data row is left untouched.
    Returns log entries for anything worth flagging during this write.
    """
    logs: List[LogEntry] = []

    shutil.copyfile(template_path, output_path)  # never touch the original
    wb = openpyxl.load_workbook(output_path)  # preserves formatting/formulas
    ws = wb.active

    cols = TIMESHEET_COLUMNS
    row_idx = TIMESHEET_FIRST_DATA_ROW

    written = 0
    for result in matched_results:
        if result.status != MatchStatus.MATCHED:
            continue
        emp = result.employee
        ws[f"{cols['eid_no']}{row_idx}"] = emp.eid_no
        ws[f"{cols['employee_name']}{row_idx}"] = emp.employee_name
        ws[f"{cols['designation']}{row_idx}"] = emp.designation
        row_idx += 1
        written += 1

    wb.save(output_path)
    wb.close()

    logs.append(LogEntry(
        issue_type="Timesheet populated",
        message=f"Wrote {written} employee row(s) into {Path(output_path).name}.",
        severity=LogSeverity.INFO,
        source_file=Path(template_path).name,
    ))
    return logs


def group_matches_by_supplier(
    results: List[MatchResult],
) -> Dict[str, List[MatchResult]]:
    """Groups MATCHED results by normalized supplier name (see normalize_supplier)."""
    grouped: Dict[str, List[MatchResult]] = {}
    for r in results:
        if r.status != MatchStatus.MATCHED or not r.supplier:
            continue
        grouped.setdefault(normalize_supplier(r.supplier), []).append(r)
    return grouped
