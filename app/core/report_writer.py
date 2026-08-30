"""
Builds the reporting workbook: New Arrivals Data Not Found,
Processing Log / Exceptions, and Processing Summary.
"""
from datetime import datetime, timezone
from typing import List

import openpyxl
from openpyxl.styles import Font, PatternFill

from app.core.config import (
    NEW_ARRIVALS_REMARK,
    NEW_ARRIVALS_SHEET_NAME,
    PROCESSING_LOG_SHEET_NAME,
    SUMMARY_SHEET_NAME,
)
from app.models.schemas import LogEntry, MatchResult, MatchStatus, ProcessingSummary

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _write_header(ws, headers: List[str]):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    ws.freeze_panes = "A2"


def build_report_workbook(
    results: List[MatchResult],
    logs: List[LogEntry],
    summary: ProcessingSummary,
    output_path: str,
) -> None:
    wb = openpyxl.Workbook()

    # --- New Arrivals sheet ---
    ws_new = wb.active
    ws_new.title = NEW_ARRIVALS_SHEET_NAME
    _write_header(ws_new, [
        "EID No.", "Extracted Employee Name", "Extracted Info", "Scanned File Name",
        "Supplier", "Date Processed", "Remarks",
    ])
    row = 2
    for r in results:
        if r.status != MatchStatus.NEW_ARRIVAL:
            continue
        ws_new.append([
            r.extracted_eid or "",
            "",  # name not derivable from EID card OCR alone
            r.notes,
            r.source_file,
            "",
            datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            NEW_ARRIVALS_REMARK,
        ])
        row += 1

    # --- Processing Log sheet ---
    ws_log = wb.create_sheet(PROCESSING_LOG_SHEET_NAME)
    _write_header(ws_log, [
        "Timestamp", "Severity", "Issue Type", "EID No.", "Source File", "Message",
    ])
    for entry in logs:
        ws_log.append([
            entry.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            entry.severity.value.upper(),
            entry.issue_type,
            entry.eid_no or "",
            entry.source_file or "",
            entry.message,
        ])

    # --- Summary sheet ---
    ws_sum = wb.create_sheet(SUMMARY_SHEET_NAME)
    ws_sum.append(["Metric", "Value"])
    for c in ws_sum[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    rows = [
        ("Total EIDs Uploaded", summary.total_eids_uploaded),
        ("Successfully Matched", summary.successfully_matched),
        ("New Arrivals / Not Found", summary.new_arrivals),
        ("Duplicate EIDs", summary.duplicate_eids),
        ("Processing Errors", summary.processing_errors),
        ("Total Suppliers", summary.total_suppliers),
        ("Total Employees Processed", summary.total_employees_processed),
        ("Timesheets Completed", summary.timesheets_completed),
        ("Report Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")),
    ]
    for label, value in rows:
        ws_sum.append([label, value])

    for ws in (ws_new, ws_log, ws_sum):
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)

    wb.save(output_path)


def build_summary(
    results: List[MatchResult],
    logs: List[LogEntry],
    total_uploaded: int,
    total_suppliers: int,
    timesheets_completed: int,
) -> ProcessingSummary:
    s = ProcessingSummary()
    s.total_eids_uploaded = total_uploaded
    s.successfully_matched = sum(1 for r in results if r.status == MatchStatus.MATCHED)
    s.new_arrivals = sum(1 for r in results if r.status == MatchStatus.NEW_ARRIVAL)
    s.duplicate_eids = sum(1 for r in results if r.status == MatchStatus.DUPLICATE)
    s.processing_errors = sum(
        1 for r in results if r.status in (MatchStatus.OCR_ERROR, MatchStatus.INVALID_FORMAT)
    )
    s.total_suppliers = total_suppliers
    s.total_employees_processed = s.successfully_matched
    s.timesheets_completed = timesheets_completed
    return s
