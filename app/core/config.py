python
"""
Populates supplier timesheet templates from matched employees.

Accepts two upload shapes for supplier templates, transparently:
  - one .xlsx file per supplier (single sheet, supplier name in H5)
  - one combined .xlsx workbook with several sheets, each sheet being
    a different supplier's timesheet (its own H5 on that sheet)
Either way, every supplier ends up with their own standalone output
file - for the combined-workbook case, all OTHER sheets are dropped
from that supplier's output copy so nobody accidentally downloads
someone else's data mixed in.

Loads each template with formatting intact, writes into a COPY (never
the original), and only ever writes into the configured data
columns/rows - it never touches the template's existing styling,
merged cells, or formulas elsewhere on the sheet.

Only EID No, Name, and Designation are written - all other columns
(time in/out, break, total hours, etc.) are left exactly as the
template already has them.
"""
import re
import shutil
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl

from app.core.config import SUPPLIER_NAME_CELL, TIMESHEET_COLUMNS, TIMESHEET_FIRST_DATA_ROW
from app.models.schemas import LogEntry, LogSeverity, MatchResult, MatchStatus


def normalize_supplier(name: str) -> str:
    """
    Case/whitespace-insensitive key for matching a master-data supplier
    name against a template's H5 supplier name. Real-world files are
    rarely byte-identical (extra spaces, different case, punctuation),
    so matching on the raw string causes a template to silently get
    zero employees written into it.
    """
    if not name:
        return ""
    return re.sub(r"\s+", " ", name).strip().casefold()


def find_supplier_sheets(filepath: str) -> List[Tuple[str, str]]:
    """
    Returns a (sheet_name, supplier_name) pair for every sheet in this
    workbook that has a supplier name in SUPPLIER_NAME_CELL. A normal
    single-supplier upload yields one pair; a combined workbook with a
    tab per supplier yields one pair per tab.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    found = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        value = ws[SUPPLIER_NAME_CELL].value
        supplier = str(value).strip() if value else ""
        if supplier:
            found.append((sheet_name, supplier))
    wb.close()
    return found


def populate_supplier_timesheet(
    template_path: str,
    sheet_name: str,
    output_path: str,
    matched_results: List[MatchResult],
) -> List[LogEntry]:
    """
    Writes matched employees' EID No, Name, and Designation into a copy
    of the given sheet within template_path. If that workbook contains
    other suppliers' sheets too (combined-file case), those are removed
    from the output copy so each supplier's file only ever contains
    their own data. Every other cell in the data row is left untouched.
    """
    logs: List[LogEntry] = []

    shutil.copyfile(template_path, output_path)  # never touch the original
    wb = openpyxl.load_workbook(output_path)  # preserves formatting/formulas

    if len(wb.sheetnames) > 1:
        for other in list(wb.sheetnames):
            if other != sheet_name:
                del wb[other]

    ws = wb[sheet_name]
    cols = TIMESHEET_COLUMNS
    row_idx = TIMESHEET_FIRST_DATA_ROW

    written = 0
    for result in matched_results:
        if result.status != MatchStatus.MATCHED:
            continue
        emp = result.employee
        ws[f"{cols['eid_no']}{row_idx}"] = emp.eid_no
        ws[f"{cols['employee_name']}{row_idx}"] = emp.employee_name
        ws[f"{cols['designation']}{row_idx}"] = emp.designation
        row_idx += 1
        written += 1

    wb.save(output_path)
    wb.close()

    logs.append(LogEntry(
        issue_type="Timesheet populated",
        message=f"Wrote {written} employee row(s) into {Path(output_path).name}.",
        severity=LogSeverity.INFO,
        source_file=Path(template_path).name,
    ))
    return logs


def group_matches_by_supplier(
    results: List[MatchResult],
) -> Dict[str, List[MatchResult]]:
    """Groups MATCHED results by normalized supplier name (see normalize_supplier)."""
    grouped: Dict[str, List[MatchResult]] = {}
    for r in results:
        if r.status != MatchStatus.MATCHED or not r.supplier:
            continue
        grouped.setdefault(normalize_supplier(r.supplier), []).append(r)
    return grouped
